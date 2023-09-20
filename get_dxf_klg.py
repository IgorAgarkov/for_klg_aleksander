# -*- coding: utf8 -*-

# import os
import pandas as pd
from shutil import copy2
from pathlib import Path
import re
import openpyxl
from datetime import datetime
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
# from openpyxl.formatting.formatting import ConditionalFormattingList

'''
Поместить скрипт в отдельную папку. Указать путь к проекту с DXF файлами, когда запросит.
Имя папки будет использоваться в названии сгенерированного файла Excel, а также будет вставлено во 2-ю колонку 2-го столбца файла Excel, как навзание заявки.
В папку со скриптом необходимо поместить файл отчёта из Компас 3D в формате .xls.
Скрипт обрабатывает этот файл, формирует перечень и сохраняет его в папке, где сам расположен. Затем ищет соответствующие DXF файлы и копирует сюда же.
Если какие-то файлы не нашлись - выдаёт сообщение об этом.
'''

def get_str_thickness(file_name):
    '''
    Функция вычленяет из имени файла числовое значение (str) по паттерну (толщина изделия).
    '''
    pattern = re.compile(
        '-(\d,?\.?\d?)[^-\d]')  # ищем: 'дефис (цифра {одна} запятая {0 или одна} точка {0 или одна} цифра {0 или одна}) не_дефис_не_цифра {одна}   - берём то что в скобках
    if len(pattern.findall(file_name)) != 0:  # если паттерн найден
        str_thickness = pattern.findall(file_name)[0].rstrip(
            '.')  # убираем точку в конце на случай, когда толщина указана в конце имени файла перед расшерением
    else:  # если паттерн не найден
        str_thickness = ''
    return str_thickness


def filling_in_excel(df, wb, first_row=1, first_col=1, ws_title=None, color_pattern=None, extra=None):
    '''
    Функция заполняет лист excel данными из датафрейма.
    В качестве параметра (color_pattern) можно передать словарь вида {'паттерн': 'цвет hex'} для
    закрашивания ячеек выбранным цветом, если паттерн подходит.
    Параметры:
    df - датафрэйм;
    wb - WorkBook-объект (не имя файла!)
    border - параметры границ ячеек
    out_path -  путь к результирующему файлу (по умолчанию ='result.xlsx');
    first_row - начальная строка заполнения, начиная с заголовка (по умолчанию =1);
    first_col - начальная колонка заполнения (по умолчанию =1);
    color_pattern - словарь вида {'паттерн': 'цвет hex'} для закрашивания ячеек цветом.
    extra - словарь вида {'ячейка': 'значение'} для вставки дополнительной информации. Ячейка - в формате 'A1' и т.п.
    '''
    # wb = openpyxl.load_workbook(template_path)                                   # открываем файл excel
    if ws_title == None:  # указываем название вкладки, куда будем записывать
        ws = wb.active
    else:
        ws = wb[ws_title]

    thin = Side(border_style="thin", color="000000")  # стиль 'thin': линии тонкие, цвет чёрный
    border = Border(top=thin, left=thin, right=thin, bottom=thin)  # параметры границ ячеек

    i = first_row  # начальная строка заполнения

    for header_no in range(0, len(df.columns)):
        ws_cell = ws.cell(row=i, column=header_no + 1)
        ws_cell.value = df.columns[header_no]
        ws_cell.border = border
        ws_cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

    i += 1  # вторая строка после заголовка имеет номер +1
    for idx, row in df.iterrows():  # итерируемся по строкам
        j = first_col  # начальная колонка заполнения
        for val in row:  # итерируемся по всем значениям кортежа row
            ws_cell = ws.cell(row=i, column=j)  # выбираем ячейку по номеру строки и колонки
            ws_cell.value = val  # присваиваем значение ячейке
            ws_cell.border = border  # устанавливаем границы ячейки
            if j != 2:  # если номер столбца не равен 2, центрируем
                ws_cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            if color_pattern != None:  # если параметр color_pattern указан, закрашиваем цветом
                for pattern, fill_color in color_pattern.items():  # итерируемся по словарю, извлекая паттерны и цвета
                    if re.search(pattern,
                                 str(ws_cell.value)) != None:  # если нашлись совпаденяия (т.е. != None), красим в цвет
                        ws_cell.fill = PatternFill("solid", fgColor=fill_color)  # параметры заполнения цветом
            j += 1
        i += 1
    if extra != None:  # итеритуемся по ключам и значениям словаря extra
        for key, value in extra.items():
            ws[key] = value  # записываем значения в ячейку


# def add_note(row):
    

path = Path('.')

for xls in path.glob('*.xls'):
    df = pd.read_excel(xls)

# паттерн для фильтрации всего, что начинается не с буквы
pattern = '^\w'  # начинается не с буквы

# убираем строки с пропусками
df.dropna(inplace=True)



# добавляем в примечания фразу "крой один, гибы зеркально" для исполнений -01
df['примечание'] = df['Обозначение'].apply(lambda x: "один крой, гибы зеркально" if x[-3:] == '-01' else "")

# убираем "-01" в обозначениях
df['Обозначение'] = df['Обозначение'].apply(lambda x: x[:-3] if x[-3:] == '-01' else x)

# фильтруем колонку, оставляя только то, что соответствует паттерну
# df = df.copy()[df['Наименование'].str.contains(pattern)]
df = df.copy()[df['Наименование'].str.contains(pattern) == False]

# получаем значения колонки с толщиной из наименования
df['Толщина, мм.'] = df['Наименование'].apply(get_str_thickness)

# переводим клонку с толщинами в числовой формат
df['Толщина, мм.'] = df['Толщина, мм.'].apply(lambda x: float(x.replace(',', '.')))

# объединяем колонки 'Обозначение' и 'Наименование'
df['Наименование'] = df['Обозначение'] + df['Наименование']

df = (
    df.groupby('Наименование', as_index=False)
        .agg({'Количество': 'sum',
              'Толщина, мм.': 'first',
              'примечание': ''.join})
    .rename(columns={'Количество': 'Кол-во, шт.'})
    )

df['№'] = df.index + 1
df['Гибка (да/нет)'] = ''
df['металл'] = ''
# df['примечание'] = ''
df = df[['№', 'Наименование', 'Кол-во, шт.', 'Толщина, мм.', 'Гибка (да/нет)', 'металл', 'примечание']]

# сегодняшняя дата
# date = datetime.today().strftime('%d.%m.%Y')

# безём название заявки из родительской директории
application = path.cwd().name

# путь к финальному файлу Excel
excel_path = path / (application + '.xlsx')

wb = openpyxl.Workbook()                   # создаём пустой файл Excel
ws = wb.active                             # лист
ws.column_dimensions['A'].width = 4.67     # устанавливаем ширину ячеек (в единицах шрифта)
ws.column_dimensions['B'].width = 63.89
ws.column_dimensions['C'].width = 8.67
ws.column_dimensions['D'].width = 9.89
ws.column_dimensions['E'].width = 9.33
ws.column_dimensions['F'].width = 9.11
ws.column_dimensions['G'].width = 25.11

ws.cell(row= 1, column = 7).value = 'Савин'                  # присваиваем значение ячейкам
ws.cell(row= 2, column = 2).value = application              # графу заявку заполняем названием родительской директории
ws.cell(row= 2, column = 2).font = Font(size=14, bold=True)  # устанавливаем размер шрифта ячейки, делаем шрифт полужирным

for i in range(1, 8):
    ws.cell(row= 3, column = i).font = Font(bold=True)       #для всей строки №3 делаем шрифт полужирным


filling_in_excel(df, wb, first_row=3, first_col=1) # заполняем Excel данными из df с помощью функции


wb.save(excel_path)                                                            # сохраняем результат
wb.close()

# путь к проекту с файлами DXF
dxf_source_path = Path(input('Введите путь к DXF файлам: '))
# пустой сет для заполнения названиями файлов DXF, чтобы найти файлы, которые присутствуют в Excel, но не нашлись в папках с DXF
dxf_list = set()
for dxf in dxf_source_path.glob('**/*.dxf'):
    if dxf.stem in df['Наименование'].values:  # если имя файла присутствует в колонке
        dxf_list.add(dxf.stem)                 # добавляем его в сет
        try:
            copy2(dxf, path / dxf.name)        # копируем в папку для заказа
        except:
            print('Не удалось скопировать файл ', dxf)
for name in df['Наименование'].values:
    if name not in dxf_list:
        print('Нет файла:', name + '.dxf')

input('Нажмите Enter для выхода')
